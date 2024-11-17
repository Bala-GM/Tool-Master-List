import sys
import os
import msoffcrypto
import io
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QLineEdit, QTableWidget,
                             QTableWidgetItem, QTabWidget, QAbstractItemView, QMessageBox,
                             QLabel, QPushButton, QDialog, QFormLayout, QDialogButtonBox)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

# Dummy data to store credentials
USERS = {
    "admin": {"password": "admin#123", "role": "admin"},
    "Process": {"password": "Pro123", "role": "process"},
    "OP": {"password": "OP", "role": "operator"}
}

def load_config(config_path):
    try:
        with open(config_path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading config file: {e}")
        return {}

class LoginPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent  # Store the reference to the parent (MainWindow)
        self.setWindowTitle("Login Page")
        self.setGeometry(250, 250, 1270, 720)

        # Create a form layout for neat alignment
        self.layout = QFormLayout()

        # Username label and input
        self.username_label = QLabel("Username:")
        self.username_input = QLineEdit()
        self.layout.addRow(self.username_label, self.username_input)

        # Password label and input
        self.password_label = QLabel("Password:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.layout.addRow(self.password_label, self.password_input)

        # Login button
        self.login_button = QPushButton("Login")
        self.login_button.clicked.connect(self.authenticate)
        self.layout.addWidget(self.login_button)

        # Set the layout for the window
        self.setLayout(self.layout)

    def authenticate(self):
        username = self.username_input.text()
        password = self.password_input.text()

        if username in USERS and USERS[username]["password"] == password:
            if self.parent is None:
                QMessageBox.warning(self, "Error", "Parent window is not set!")
                return

            self.parent.user_role = USERS[username]["role"]
            self.parent.show_role_specific_page()
        else:
            QMessageBox.warning(self, "Login Failed", "Invalid username or password")

class ExcelViewerWithHomePage(QWidget):
    def __init__(self, file_path, password=None, hidden_sheets=None):
        super().__init__()
        self.file_path = file_path  # Save the file path for later use
        self.setWindowTitle("Excel Viewer with Home and Search Navigation")
        self.setGeometry(0, 0, 1920, 1080)  # Full-screen size
        self.hidden_sheets = hidden_sheets or []

        try:
            decrypted_file = io.BytesIO()
            with open(file_path, "rb") as f:
                encrypted = msoffcrypto.OfficeFile(f)
                encrypted.load_key(password=password)
                encrypted.decrypt(decrypted_file)

            decrypted_file.seek(0)
            self.workbook = load_workbook(decrypted_file, data_only=True, keep_vba=True)

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load or decrypt Excel file: {e}")
            return

        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Search...")
        self.search_bar.textChanged.connect(self.search)

        self.tab_widget = QTabWidget()
        self.home_table = QTableWidget()
        self.home_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.home_table.cellClicked.connect(self.navigate_to_sheet)
        self.tab_widget.addTab(self.home_table, "Home")

        self.tables = {}
        self.load_sheets()

        layout = QVBoxLayout()
        layout.addWidget(self.search_bar)
        layout.addWidget(self.tab_widget)
        self.setLayout(layout)

    def load_sheets(self):
        try:
            for sheet_name in self.workbook.sheetnames:
                if sheet_name in self.hidden_sheets:
                    continue

                sheet = self.workbook[sheet_name]
                table_widget = QTableWidget()
                table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
                table_widget.setRowCount(sheet.max_row)
                table_widget.setColumnCount(sheet.max_column)

                self.load_data(sheet, table_widget)
                self.tables[sheet_name] = table_widget
                self.tab_widget.addTab(table_widget, sheet_name)

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load sheets: {e}")

    def load_data(self, sheet, table_widget):
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    item = QTableWidgetItem(str(cell.value))
                    item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(cell.row - 1, cell.column - 1, item)

            for merged_range in sheet.merged_cells.ranges:
                start_row = merged_range.min_row - 1
                end_row = merged_range.max_row - 1
                start_col = merged_range.min_col - 1
                end_col = merged_range.max_col - 1
                table_widget.setSpan(start_row, start_col, end_row - start_row + 1, end_col - start_col + 1)

            # Adjust column widths based on content
            table_widget.resizeColumnsToContents()

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load sheet data: {e}")

    def search(self, text):
        self.home_table.clear()
        self.home_table.setRowCount(0)
        self.home_table.setColumnCount(2)
        self.home_table.setHorizontalHeaderLabels(["Sheet", "Content"])

        for sheet_name, table_widget in self.tables.items():
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    if item is not None:
                        if text.lower() in item.text().lower():
                            result_row = self.home_table.rowCount()
                            self.home_table.insertRow(result_row)
                            self.home_table.setItem(result_row, 0, QTableWidgetItem(sheet_name))
                            self.home_table.setItem(result_row, 1, QTableWidgetItem(item.text()))

                            # Highlight only the matching item
                            item.setBackground(Qt.yellow)
                        else:
                            # Reset non-matching cells to normal
                            item.setBackground(Qt.transparent)

    def navigate_to_sheet(self, row, column):
        sheet_name = self.home_table.item(row, 0).text()
        content = self.home_table.item(row, 1).text()
        self.tab_widget.setCurrentWidget(self.tables[sheet_name])

        for r in range(self.tables[sheet_name].rowCount()):
            for c in range(self.tables[sheet_name].columnCount()):
                item = self.tables[sheet_name].item(r, c)
                if item and item.text() == content:
                    self.tables[sheet_name].setCurrentCell(r, c)
                    return

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Tool Master")
        self.setGeometry(250, 250, 1270, 720)

        self.layout = QVBoxLayout()

        # Pass self as the parent to LoginPage
        self.login_page = LoginPage(self)
        self.layout.addWidget(self.login_page)

        self.setLayout(self.layout)

    def show_role_specific_page(self):
        self.login_page.hide()

        if self.user_role == "operator":
            self.show_op_page()
        else:
            QMessageBox.information(self, "Info", "This role is not implemented yet.")

    def show_op_page(self):
        config_data = load_config("config.json")
        op_file_path = config_data.get("op_file_path", "")
        workbook_password = config_data.get("workbook_password", "")
        hidden_sheets = config_data.get("hidden_sheets", [])

        if op_file_path:
            try:
                self.viewer = ExcelViewerWithHomePage(file_path=op_file_path, password=workbook_password, hidden_sheets=hidden_sheets)
                self.layout.addWidget(self.viewer)
                self.viewer.show()

            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to load Excel file: {e}")
        else:
            QMessageBox.warning(self, "Error", "Excel file path is missing in the config.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

#pyinstaller --onefile ToolMaster.py
